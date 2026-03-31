"""
EDGAR SAP ECC Prospector — FASB DISE Solution
================================================
Pulls 10-K filings from SEC EDGAR full-text search (EFTS) for companies
disclosing SAP ECC, scores each company by signal strength, deduplicates,
and writes a tiered prospect spreadsheet.

Usage:
    pip install requests openpyxl
    python edgar_sap_ecc_prospector.py

Output:
    sap_ecc_prospects_YYYYMMDD.xlsx  (in current directory)

Requirements: Python 3.8+, internet access to efts.sec.gov
"""

import requests
import time
import re
import json
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

START_DATE = "2022-01-01"
END_DATE   = "2025-12-31"
PAGE_SIZE  = 100          # max per EFTS request
MAX_PAGES  = 10           # cap at 1,000 results per query
RATE_LIMIT = 0.5          # seconds between requests (be polite to SEC)

EFTS_BASE = "https://efts.sec.gov/LATEST/search-index"
EDGAR_BASE = "https://www.sec.gov/cgi-bin/browse-edgar"

HEADERS = {
    "User-Agent": "FASB-DISE-Research ujjwal@truffles.ai",
    "Accept-Encoding": "gzip, deflate",
}

# ─────────────────────────────────────────────
# SEARCH QUERIES  (name, query_string, tier_boost)
# ─────────────────────────────────────────────

# Each query is a simple phrase — EFTS handles these reliably via GET
QUERIES = [
    {"name": "SAP ECC direct",          "q": '"SAP ECC"',                    "tier_boost": 0,  "signal": "direct_ecc"},
    {"name": "ERP Central Component",   "q": '"ERP Central Component"',      "tier_boost": 0,  "signal": "direct_ecc"},
    {"name": "ECC 6.0",                 "q": '"ECC 6.0"',                    "tier_boost": 0,  "signal": "direct_ecc"},
    {"name": "S/4HANA migration",       "q": '"S/4HANA migration"',          "tier_boost": 10, "signal": "migration_active"},
    {"name": "RISE with SAP",           "q": '"RISE with SAP"',              "tier_boost": 10, "signal": "migration_active"},
    {"name": "Rimini Street SAP",       "q": '"Rimini Street" "SAP"',        "tier_boost": 8,  "signal": "rimini_street"},
    {"name": "SAP end of support",      "q": '"SAP" "end of support"',       "tier_boost": 12, "signal": "risk_factor_ecc"},
    {"name": "SAP end of maintenance",  "q": '"SAP" "end of maintenance"',   "tier_boost": 12, "signal": "risk_factor_ecc"},
    {"name": "Legacy SAP migration",    "q": '"legacy" "SAP" "migration"',   "tier_boost": 6,  "signal": "erp_upgrade_risk"},
    {"name": "SAP NetWeaver",           "q": '"SAP NetWeaver"',              "tier_boost": 4,  "signal": "netweaver"},
    {"name": "SAP Business Suite",      "q": '"SAP Business Suite"',         "tier_boost": 4,  "signal": "netweaver"},
    {"name": "SAP financial close",     "q": '"SAP" "financial close"',      "tier_boost": 15, "signal": "fasb_adjacent"},
    {"name": "SAP FASB disclosure",     "q": '"SAP" "FASB" "ECC"',          "tier_boost": 15, "signal": "fasb_adjacent"},
]

# SIC codes with highest ECC probability (used for industry labeling)
SIC_INDUSTRY_MAP = {
    "1311": "Oil & Gas Extraction",
    "1381": "Oil & Gas Services",
    "1382": "Oil & Gas Field Services",
    "2000": "Food Manufacturing",
    "2090": "Food & Kindred Products",
    "2110": "Tobacco Products",
    "2800": "Chemicals",
    "2810": "Industrial Chemicals",
    "2820": "Plastics & Synthetics",
    "2830": "Pharmaceuticals",
    "2833": "Pharmaceutical Preparations",
    "2836": "Biological Products",
    "3310": "Steel & Iron",
    "3559": "Industrial Machinery",
    "3560": "General Industrial Machinery",
    "3562": "Ball & Roller Bearings",
    "3670": "Electronic Components",
    "3672": "Printed Circuit Boards",
    "3674": "Semiconductors",
    "3711": "Motor Vehicles",
    "3714": "Motor Vehicle Parts",
    "3716": "Motor Homes",
    "3720": "Aircraft",
    "3724": "Aircraft Engine Parts",
    "3728": "Aircraft Parts",
    "4911": "Electric Services",
    "4931": "Electric & Gas Services",
    "4941": "Water Supply",
    "5311": "Department Stores",
    "5331": "Variety Stores",
    "5411": "Grocery Stores",
}

HIGH_VALUE_SICS = set(SIC_INDUSTRY_MAP.keys())


# ─────────────────────────────────────────────
# EDGAR EFTS FETCHER
# ─────────────────────────────────────────────

def fetch_efts_page(query: str, from_offset: int) -> dict:
    """
    Fetch one page of EFTS results via GET.
    Uses the EDGAR full-text search API with simple phrase query.
    """
    params = {
        "q": query,
        "forms": "10-K",
        "dateRange": "custom",
        "startdt": START_DATE,
        "enddt": END_DATE,
        "from": from_offset,
        "size": PAGE_SIZE,
    }
    resp = requests.get(EFTS_BASE, params=params, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return resp.json()


def fetch_all_hits(query_cfg: dict) -> list[dict]:
    """Page through EFTS and return all hits for a query."""
    query_str = query_cfg["q"]
    print(f"  Searching: {query_cfg['name']} ...", end="", flush=True)

    all_hits = []
    offset = 0

    for page in range(MAX_PAGES):
        try:
            data = fetch_efts_page(query_str, offset)
        except requests.RequestException as e:
            print(f"\n  [WARN] Request failed: {e}")
            break

        hits = data.get("hits", {}).get("hits", [])
        total = data.get("hits", {}).get("total", {}).get("value", 0)

        if not hits:
            break

        for h in hits:
            src = h.get("_source", {})
            # Normalize EFTS field names
            display = (src.get("display_names") or [""])[0]
            cik = (src.get("ciks") or [""])[0]
            sic = (src.get("sics") or [""])[0]
            normalized = {
                "entity_name": display.split("(CIK")[0].strip().rstrip(")").strip() if display else "Unknown",
                "entity_id": cik,
                "sic": sic,
                "file_num": (src.get("file_num") or [""])[0],
                "file_date": src.get("file_date", ""),
                "period_of_report": src.get("period_ending", ""),
                "accession_no": src.get("adsh", ""),
                "_query_name": query_cfg["name"],
                "_signal": query_cfg["signal"],
                "_tier_boost": query_cfg["tier_boost"],
            }
            all_hits.append(normalized)

        offset += PAGE_SIZE
        if offset >= total or offset >= MAX_PAGES * PAGE_SIZE:
            break

        time.sleep(RATE_LIMIT)

    print(f" {len(all_hits)} filings found")
    return all_hits


# ─────────────────────────────────────────────
# SCORING ENGINE
# ─────────────────────────────────────────────

def score_company(signals: list[str], sic: str, latest_filing: str) -> dict:
    score = 0
    breakdown = []

    signal_scores = {
        "direct_ecc":        30,
        "risk_factor_ecc":   25,
        "migration_active":  20,
        "fasb_adjacent":     18,
        "rimini_street":     15,
        "erp_upgrade_risk":  10,
        "netweaver":          8,
    }

    seen = set()
    for sig in signals:
        if sig not in seen:
            pts = signal_scores.get(sig, 0)
            score += pts
            breakdown.append(f"{sig}: +{pts}")
            seen.add(sig)

    distinct = len(seen)
    if distinct >= 3:
        score += 10
        breakdown.append("multi-signal corroboration: +10")
    elif distinct == 2:
        score += 5
        breakdown.append("dual-signal corroboration: +5")

    if sic in HIGH_VALUE_SICS:
        score += 8
        breakdown.append(f"priority SIC {sic}: +8")

    try:
        filing_date = datetime.strptime(latest_filing[:10], "%Y-%m-%d")
        months_ago = (datetime.now() - filing_date).days / 30
        if months_ago <= 6:
            score += 7
            breakdown.append("filed <6 months ago: +7")
        elif months_ago <= 18:
            score += 4
            breakdown.append("filed <18 months ago: +4")
    except Exception:
        pass

    score = min(score, 100)
    if score >= 60:
        tier = "1 — High Priority"
    elif score >= 35:
        tier = "2 — Medium Priority"
    else:
        tier = "3 — Monitor"

    return {"score": score, "tier": tier, "breakdown": " | ".join(breakdown), "signals_count": distinct}


# ─────────────────────────────────────────────
# DEDUPLICATION & AGGREGATION
# ─────────────────────────────────────────────

def deduplicate(all_hits: list[dict]) -> list[dict]:
    companies: dict[str, dict] = {}

    for hit in all_hits:
        key = hit.get("entity_id") or hit.get("entity_name", "").upper().strip()
        if not key:
            continue

        if key not in companies:
            companies[key] = {
                "entity_name": hit.get("entity_name", "Unknown"),
                "entity_id": hit.get("entity_id", ""),
                "file_num": hit.get("file_num", ""),
                "sic": hit.get("sic", ""),
                "signals": [],
                "query_names": [],
                "latest_filing": "",
                "filing_count": 0,
                "accession_nos": [],
            }

        co = companies[key]
        co["signals"].append(hit.get("_signal", ""))
        co["query_names"].append(hit.get("_query_name", ""))
        co["filing_count"] += 1

        acc = hit.get("accession_no", "")
        if acc and acc not in co["accession_nos"]:
            co["accession_nos"].append(acc)

        fd = hit.get("file_date", "")
        if fd > co["latest_filing"]:
            co["latest_filing"] = fd

    return list(companies.values())


def enrich_companies(companies: list[dict]) -> list[dict]:
    enriched = []

    for co in companies:
        sic = co.get("sic", "")
        scored = score_company(co["signals"], sic, co["latest_filing"])
        industry = SIC_INDUSTRY_MAP.get(sic, f"SIC {sic}" if sic else "")

        cik = co.get("entity_id", "").lstrip("0")
        edgar_link = f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={cik}&type=10-K&dateb=&owner=include&count=10" if cik else ""

        unique_signals = sorted(set(co["signals"]))
        unique_queries = sorted(set(co["query_names"]))

        dise_notes = []
        if "direct_ecc" in unique_signals:
            dise_notes.append("Explicitly names SAP ECC in 10-K")
        if "risk_factor_ecc" in unique_signals:
            dise_notes.append("ECC flagged as material risk")
        if "fasb_adjacent" in unique_signals:
            dise_notes.append("SAP + financial reporting disclosure gap identified")
        if "rimini_street" in unique_signals:
            dise_notes.append("Delayed migration via Rimini Street — active ECC pain")
        if "migration_active" in unique_signals:
            dise_notes.append("S/4HANA migration in progress — ECC still live")

        enriched.append({
            "Company Name": co["entity_name"],
            "CIK": co.get("entity_id", ""),
            "SIC Code": sic,
            "Industry": industry,
            "Tier": scored["tier"],
            "Score (0–100)": scored["score"],
            "Signals Matched": scored["signals_count"],
            "Signal Details": " | ".join(unique_signals),
            "Search Queries Hit": " | ".join(unique_queries),
            "10-K Filings Found": co["filing_count"],
            "Latest Filing Date": co["latest_filing"][:10] if co["latest_filing"] else "",
            "DISE Opportunity Notes": "; ".join(dise_notes) if dise_notes else "Review filing text",
            "Score Breakdown": scored["breakdown"],
            "EDGAR 10-K Link": edgar_link,
            "File Number": co.get("file_num", ""),
        })

    enriched.sort(key=lambda x: (-x["Score (0–100)"], x["Company Name"]))
    return enriched


# ─────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────

C_HEADER_BG = "1F3864"
C_HEADER_FG = "FFFFFF"
C_T1_BG = "E8F5E9"
C_T1_ACCENT = "2E7D32"
C_T2_BG = "FFF8E1"
C_T2_ACCENT = "F57F17"
C_T3_BG = "F5F5F5"
C_T3_ACCENT = "757575"
C_ALT_ROW = "F9FAFB"
C_DIVIDER = "E0E0E0"
C_LINK = "1565C0"

TIER_COLORS = {"1": (C_T1_BG, C_T1_ACCENT), "2": (C_T2_BG, C_T2_ACCENT), "3": (C_T3_BG, C_T3_ACCENT)}

def thin_border():
    s = Side(style="thin", color=C_DIVIDER)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(): return Font(name="Arial", bold=True, color=C_HEADER_FG, size=10)
def body_font(bold=False, color="000000", size=9): return Font(name="Arial", bold=bold, color=color, size=size)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def fill(hex_color): return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def write_cover_sheet(wb, total, by_tier, run_date):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:H3")
    ws["B2"] = "SAP ECC Prospect Intelligence — FASB DISE Solution"
    ws["B2"].font = Font(name="Arial", bold=True, size=16, color=C_HEADER_BG)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("B4:H4")
    ws["B4"] = f"SEC EDGAR 10-K Analysis  |  Filings: {START_DATE} – {END_DATE}  |  Generated: {run_date}"
    ws["B4"].font = body_font(color="666666", size=9)
    ws["B4"].alignment = left()

    stats = [("Total Companies", total, C_HEADER_BG), ("Tier 1 — High Priority", by_tier.get("1", 0), C_T1_ACCENT),
             ("Tier 2 — Medium", by_tier.get("2", 0), "B45309"), ("Tier 3 — Monitor", by_tier.get("3", 0), C_T3_ACCENT)]
    for i, (label, val, color) in enumerate(stats):
        col = [2, 4, 6, 8][i]
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 1)
        ws.cell(row=6, column=col, value=val).font = Font(name="Arial", bold=True, size=22, color=color)
        ws.cell(row=7, column=col, value=label).font = body_font(color="555555", size=9)
        ws.cell(row=6, column=col).alignment = center()
        ws.cell(row=7, column=col).alignment = center()


def write_prospects_sheet(wb, rows):
    ws = wb.create_sheet("All Prospects")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    if not rows: return

    columns = list(rows[0].keys())
    col_widths = {"Company Name": 28, "CIK": 10, "SIC Code": 8, "Industry": 20, "Tier": 20,
                  "Score (0–100)": 12, "Signals Matched": 12, "Signal Details": 35,
                  "Search Queries Hit": 35, "10-K Filings Found": 12, "Latest Filing Date": 14,
                  "DISE Opportunity Notes": 45, "Score Breakdown": 50, "EDGAR 10-K Link": 14, "File Number": 12}

    for j, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = header_font()
        c.fill = fill(C_HEADER_BG)
        c.alignment = center()
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(j)].width = col_widths.get(col, 14)

    for i, row in enumerate(rows, 2):
        tier_num = row["Tier"][:1]
        bg, fg = TIER_COLORS.get(tier_num, (C_ALT_ROW, "000000"))
        row_bg = bg if i % 2 == 0 else C_ALT_ROW
        for j, col in enumerate(columns, 1):
            val = row[col]
            c = ws.cell(row=i, column=j, value=val)
            c.border = thin_border()
            c.alignment = left()
            if col == "Company Name":
                c.font = body_font(bold=True, size=9)
            elif col == "Tier":
                c.font = body_font(bold=True, color=fg, size=9)
                c.fill = fill(bg)
                c.alignment = center()
                continue
            elif col == "EDGAR 10-K Link" and val:
                c.value = "View Filings"
                c.font = Font(name="Arial", size=9, color=C_LINK, underline="single")
                c.hyperlink = val
                c.alignment = center()
                continue
            else:
                c.font = body_font(size=9)
            c.fill = fill(row_bg)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    score_col = columns.index("Score (0–100)") + 1
    ws.conditional_formatting.add(
        f"{get_column_letter(score_col)}2:{get_column_letter(score_col)}{len(rows)+1}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F7C1C1",
                       mid_type="num", mid_value=50, mid_color="FAEEDA",
                       end_type="num", end_value=100, end_color="C0DD97"))


def write_tier_sheet(wb, rows, tier, sheet_name):
    tier_rows = [r for r in rows if r["Tier"].startswith(tier)]
    if not tier_rows: return

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    show_cols = ["Company Name", "Industry", "Score (0–100)", "Signals Matched",
                 "DISE Opportunity Notes", "Latest Filing Date", "EDGAR 10-K Link"]
    bg, fg = TIER_COLORS.get(tier, (C_ALT_ROW, "000000"))
    widths = [30, 22, 12, 12, 50, 14, 14]

    for j, (col, w) in enumerate(zip(show_cols, widths), 1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = fill(fg)
        c.alignment = center()
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(j)].width = w

    for i, row in enumerate(tier_rows, 2):
        row_bg = bg if i % 2 == 0 else C_ALT_ROW
        for j, col in enumerate(show_cols, 1):
            val = row[col]
            c = ws.cell(row=i, column=j, value=val)
            c.border = thin_border()
            c.font = body_font(size=9, bold=(col == "Company Name"))
            c.fill = fill(row_bg)
            c.alignment = center() if col in ["Score (0–100)", "Signals Matched", "Latest Filing Date"] else left()
            if col == "EDGAR 10-K Link" and val:
                c.value = "View"
                c.font = Font(name="Arial", size=9, color=C_LINK, underline="single")
                c.hyperlink = val
                c.alignment = center()
    ws.auto_filter.ref = f"A1:{get_column_letter(len(show_cols))}1"


def write_raw_signals_sheet(wb, all_hits):
    ws = wb.create_sheet("Raw Filing Hits")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    headers = ["Entity Name", "Filing Date", "Period", "Signal", "Query", "Accession No.", "Tier Boost"]
    widths = [32, 14, 12, 22, 35, 22, 10]

    for j, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = header_font()
        c.fill = fill(C_HEADER_BG)
        c.alignment = center()
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(j)].width = w

    for i, hit in enumerate(all_hits[:2000], 2):
        vals = [hit.get("entity_name", ""), hit.get("file_date", "")[:10] if hit.get("file_date") else "",
                hit.get("period_of_report", "")[:10] if hit.get("period_of_report") else "",
                hit.get("_signal", ""), hit.get("_query_name", ""), hit.get("accession_no", ""), hit.get("_tier_boost", 0)]
        row_bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = body_font(size=9)
            c.fill = fill(row_bg)
            c.border = thin_border()
            c.alignment = left()
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def run(use_sample: bool = False):
    run_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    filename = f"sap_ecc_prospects_{datetime.now().strftime('%Y%m%d')}.xlsx"

    print("=" * 58)
    print("  EDGAR SAP ECC Prospector — FASB DISE Solution")
    print("=" * 58)

    if use_sample:
        print("\n[INFO] Running in SAMPLE mode (no live EDGAR calls)")
        all_hits = SAMPLE_HITS
    else:
        print(f"\nFetching 10-K filings from EDGAR: {START_DATE} -> {END_DATE}")
        all_hits = []
        for qcfg in QUERIES:
            hits = fetch_all_hits(qcfg)
            all_hits.extend(hits)
            time.sleep(RATE_LIMIT)

    print(f"\nTotal raw filing hits: {len(all_hits)}")

    companies = deduplicate(all_hits)
    print(f"Unique companies after dedup: {len(companies)}")
    enriched = enrich_companies(companies)

    by_tier = defaultdict(int)
    for r in enriched:
        by_tier[r["Tier"][:1]] += 1
    print(f"\nTier 1 (High Priority): {by_tier['1']}")
    print(f"Tier 2 (Medium):        {by_tier['2']}")
    print(f"Tier 3 (Monitor):       {by_tier['3']}")

    print(f"\nWriting {filename} ...")
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    write_cover_sheet(wb, len(enriched), by_tier, run_date)
    write_prospects_sheet(wb, enriched)
    write_tier_sheet(wb, enriched, "1", "Tier 1 — High Priority")
    write_tier_sheet(wb, enriched, "2", "Tier 2 — Medium")
    write_tier_sheet(wb, enriched, "3", "Tier 3 — Monitor")
    write_raw_signals_sheet(wb, all_hits)

    wb.save(filename)
    print(f"\nDone -> {filename}")
    return filename


# Sample data for offline testing
SAMPLE_HITS = [
    {"entity_name":"Interpublic Group of Companies Inc","entity_id":"0000051644","sic":"7311","file_num":"001-06431","file_date":"2024-02-21","period_of_report":"2023-12-31","accession_no":"0000051644-24-000010","_query_name":"Core ECC terms","_signal":"direct_ecc","_tier_boost":0},
    {"entity_name":"Interpublic Group of Companies Inc","entity_id":"0000051644","sic":"7311","file_num":"001-06431","file_date":"2024-02-21","period_of_report":"2023-12-31","accession_no":"0000051644-24-000010","_query_name":"S/4HANA migration underway","_signal":"migration_active","_tier_boost":10},
    {"entity_name":"Interpublic Group of Companies Inc","entity_id":"0000051644","sic":"7311","file_num":"001-06431","file_date":"2024-02-21","period_of_report":"2023-12-31","accession_no":"0000051644-24-000010","_query_name":"FASB-adjacent SAP disclosure","_signal":"fasb_adjacent","_tier_boost":15},
    {"entity_name":"Eastman Chemical Company","entity_id":"0000915389","sic":"2820","file_num":"001-12626","file_date":"2024-02-16","period_of_report":"2023-12-31","accession_no":"0000915389-24-000012","_query_name":"Core ECC terms","_signal":"direct_ecc","_tier_boost":0},
    {"entity_name":"Eastman Chemical Company","entity_id":"0000915389","sic":"2820","file_num":"001-12626","file_date":"2024-02-16","period_of_report":"2023-12-31","accession_no":"0000915389-24-000012","_query_name":"ECC end-of-support risk factor","_signal":"risk_factor_ecc","_tier_boost":12},
    {"entity_name":"Eastman Chemical Company","entity_id":"0000915389","sic":"2820","file_num":"001-12626","file_date":"2024-02-16","period_of_report":"2023-12-31","accession_no":"0000915389-24-000012","_query_name":"FASB-adjacent SAP disclosure","_signal":"fasb_adjacent","_tier_boost":15},
    {"entity_name":"Kennametal Inc","entity_id":"0000055242","sic":"3559","file_num":"001-05765","file_date":"2024-08-13","period_of_report":"2024-06-30","accession_no":"0000055242-24-000032","_query_name":"Core ECC terms","_signal":"direct_ecc","_tier_boost":0},
    {"entity_name":"Kennametal Inc","entity_id":"0000055242","sic":"3559","file_num":"001-05765","file_date":"2024-08-13","period_of_report":"2024-06-30","accession_no":"0000055242-24-000032","_query_name":"S/4HANA migration underway","_signal":"migration_active","_tier_boost":10},
    {"entity_name":"Kennametal Inc","entity_id":"0000055242","sic":"3559","file_num":"001-05765","file_date":"2024-08-13","period_of_report":"2024-06-30","accession_no":"0000055242-24-000032","_query_name":"ECC end-of-support risk factor","_signal":"risk_factor_ecc","_tier_boost":12},
    {"entity_name":"Kennametal Inc","entity_id":"0000055242","sic":"3559","file_num":"001-05765","file_date":"2024-08-13","period_of_report":"2024-06-30","accession_no":"0000055242-24-000032","_query_name":"Third-party SAP support (ECC delay)","_signal":"rimini_street","_tier_boost":8},
]


if __name__ == "__main__":
    import sys
    use_sample = "--sample" in sys.argv or "-s" in sys.argv
    run(use_sample=use_sample)
