#!/usr/bin/env python3
"""
BE Technology — GL Intelligence Platform
Offline Mapping Agent — Runs against Excel data (no BigQuery required)
─────────────────────────────────────────────────────────────────────
Reads GL accounts from DISE_Manual_Work_Demo.xlsx and classifies each
into DISE categories using Claude. Uses pre-mapped accounts from
DISE_Mapping_Toolkit.xlsx as the similarity reference library.

Outputs results to dise_mapping_results.json and dise_mapping_results.xlsx

Usage:
  python3 run_offline_agent.py                    # classify 10 accounts (quick test)
  python3 run_offline_agent.py --batch 50         # classify 50 accounts
  python3 run_offline_agent.py --batch all        # classify all 502 accounts
  python3 run_offline_agent.py --dry-run          # show prompts without calling Claude

Environment:
  ANTHROPIC_API_KEY=sk-ant-...
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import logging
from datetime import datetime, timezone
from difflib import SequenceMatcher

import anthropic

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
)
log = logging.getLogger('offline-agent')

# ── Configuration ──────────────────────────────────────────────
MODEL = os.environ.get('CLAUDE_MODEL', 'claude-sonnet-4-20250514')
API_DELAY = float(os.environ.get('API_DELAY_SECONDS', '0.3'))

VALID_CATEGORIES = [
    'Purchases of inventory',
    'Employee compensation',
    'Depreciation',
    'Intangible asset amortization',
    'Other expenses',
]
VALID_CAPTIONS = ['COGS', 'SG&A', 'R&D', 'Other income/expense']

SYSTEM_PROMPT = """You are the GL Mapping Agent for BE Technology's GL Intelligence Platform.

Your job is to classify General Ledger accounts into the five natural expense categories
required by ASU 2024-03 (DISE) under ASC 220-40.

THE FIVE DISE CATEGORIES:

1. Purchases of inventory — ASC 220-40-50-6(b)
   Includes: cost of goods purchased for resale, raw materials, direct materials, freight-in
   Excludes: labour costs, overhead, depreciation of manufacturing equipment
   Caption: almost always COGS

2. Employee compensation — ASC 220-40-50-6(a)
   Includes: salaries, wages, bonuses, commissions, payroll taxes, benefits, pension,
             stock-based comp, severance, recruitment, training, T&E for employees
   Excludes: payments to contractors/vendors (those go to Other expenses)
   Caption: COGS if production labour, SG&A if selling/admin, R&D if R&D staff

3. Depreciation — ASC 220-40-50-6(c)
   Includes: depreciation of TANGIBLE fixed assets — buildings, machinery, equipment,
             vehicles, computers/hardware, furniture, leasehold improvements, ROU assets (ASC 842)
   Excludes: amortization of INTANGIBLE assets
   CRITICAL: Computer HARDWARE = Depreciation. Computer SOFTWARE = Amortization.
   Caption: COGS if production asset, SG&A if office/admin

4. Intangible asset amortization — ASC 220-40-50-6(d)
   Includes: patents, trademarks, customer relationships, trade names, non-competes,
             developed technology, computer software (ASC 350-40), goodwill impairment
   Caption: almost always SG&A

5. Other expenses — ASC 220-40-50-6(e)
   Includes: everything else — rent, utilities, insurance, professional fees, marketing,
             advertising, R&D costs, bad debt, FX losses, bank fees, contractor services,
             temporary staffing, consulting
   Caption: depends on function

EDGE CASES:
- Operating lease payments → Other expenses (not depreciation)
- ROU asset depreciation → Depreciation (ASC 842 creates a tangible asset)
- Contractor/outsourced services → Other expenses
- Software development labor by employees → Employee compensation
- Capitalized software amortization → Intangible asset amortization
- Freight-in on inventory → Purchases of inventory

VALID CAPTIONS: COGS | SG&A | R&D | Other income/expense

CONFIDENCE:
- HIGH (0.85-1.0): Unambiguous match, supported by similar accounts
- MEDIUM (0.60-0.84): Likely match but some ambiguity
- LOW (0.0-0.59): Ambiguous — human must investigate

Respond ONLY with valid JSON:
{
  "suggested_category": "<exact category name>",
  "suggested_caption": "<COGS | SG&A | R&D | Other income/expense>",
  "suggested_citation": "<ASC citation>",
  "confidence_score": <float 0.0-1.0>,
  "confidence_label": "<HIGH | MEDIUM | LOW>",
  "draft_reasoning": "<3+ sentence explanation for the Controller>"
}"""


# ══════════════════════════════════════════════════════════════
# DATA LOADING — Excel
# ══════════════════════════════════════════════════════════════

def load_gl_accounts(path: str = 'DISE_Manual_Work_Demo.xlsx') -> list[dict]:
    """Load 502 GL accounts from the Manual Work Demo spreadsheet."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[1]  # "GL Account Master" sheet

    accounts = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue
        accounts.append({
            'gl_account': str(vals[0]).strip(),
            'description': str(vals[1] or '').strip(),
            'functional_area': str(vals[2] or '').strip(),
            'cost_element_type': str(vals[3] or '').strip(),
            'posting_amount': float(vals[4] or 0),
            'fy2022_balance': float(vals[5] or 0),
        })
    wb.close()
    return accounts


def load_reference_mappings(path: str = 'DISE_Mapping_Toolkit.xlsx') -> list[dict]:
    """Load 15 pre-mapped accounts as the similarity reference library."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[2]  # "Category Mapping" sheet

    mapped = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        vals = list(row)
        if not vals[0] or not vals[3]:
            continue
        # Normalize category names to match our valid list
        raw_cat = str(vals[3]).strip()
        cat = _normalize_category(raw_cat)
        # Normalize caption
        raw_caption = str(vals[2] or '').strip()
        caption = _normalize_caption(raw_caption)

        mapped.append({
            'gl_account': str(vals[0]).strip(),
            'description': str(vals[1] or '').strip(),
            'expense_caption': caption,
            'dise_category': cat,
            'asc_citation': _extract_citation(str(vals[4] or '')),
            'rationale': str(vals[4] or '').strip(),
        })
    wb.close()
    return mapped


def _normalize_category(raw: str) -> str:
    """Map spreadsheet category names to the exact valid names."""
    mapping = {
        'purchases of inventory': 'Purchases of inventory',
        'employee compensation': 'Employee compensation',
        'depreciation': 'Depreciation',
        'intangible asset amortization': 'Intangible asset amortization',
        'other expenses': 'Other expenses',
        'other': 'Other expenses',
    }
    return mapping.get(raw.lower().strip(), 'Other expenses')


def _normalize_caption(raw: str) -> str:
    """Map spreadsheet captions to valid values."""
    r = raw.lower().strip()
    if 'cogs' in r or 'cost of' in r or 'products sold' in r:
        return 'COGS'
    if 'sg&a' in r or 'general' in r or 'admin' in r or 'selling' in r:
        return 'SG&A'
    if 'r&d' in r or 'research' in r:
        return 'R&D'
    return 'SG&A'  # default


def _extract_citation(text: str) -> str:
    """Extract ASC citation from rationale text."""
    match = re.search(r'ASC\s+[\d-]+(?:\s*[-–]\s*[\d]+)*(?:\([a-z]\))?', text)
    return match.group(0) if match else 'ASC 220-40-50-6(e)'


# ══════════════════════════════════════════════════════════════
# SIMILARITY — Word-overlap Jaccard (same logic as BigQuery version)
# ══════════════════════════════════════════════════════════════

def _tokenize(text: str) -> set[str]:
    """Split description into lowercase word tokens > 2 chars."""
    words = re.split(r'[-/&,.\s]+', text.lower())
    return {w for w in words if len(w) > 2}


def find_similar(description: str, reference: list[dict], top_n: int = 5) -> list[dict]:
    """Find most similar pre-mapped accounts using Jaccard word overlap."""
    query_tokens = _tokenize(description)
    if not query_tokens:
        return []

    scored = []
    for ref in reference:
        ref_tokens = _tokenize(ref['description'])
        if not ref_tokens:
            continue
        intersection = len(query_tokens & ref_tokens)
        union = len(query_tokens | ref_tokens)
        score = intersection / union if union > 0 else 0
        if score > 0:
            scored.append({**ref, 'similarity_score': round(score, 3)})

    scored.sort(key=lambda x: x['similarity_score'], reverse=True)
    return scored[:top_n]


# ══════════════════════════════════════════════════════════════
# CLAUDE API
# ══════════════════════════════════════════════════════════════

def build_prompt(account: dict, similar: list[dict]) -> str:
    """Build the classification prompt for one account."""
    similar_text = ""
    if similar:
        similar_text = "\n\nSIMILAR PRE-MAPPED ACCOUNTS (use as reference):\n"
        for i, s in enumerate(similar, 1):
            similar_text += (
                f"{i}. GL {s['gl_account']} — \"{s['description']}\"\n"
                f"   Category: {s['dise_category']} | Caption: {s['expense_caption']}\n"
                f"   Citation: {s.get('asc_citation', 'N/A')} | Similarity: {s['similarity_score']:.3f}\n"
            )
    else:
        similar_text = "\n\nNo similar pre-mapped accounts found. Classify from first principles.\n"

    return f"""Classify this GL account into one of the five DISE categories.

ACCOUNT TO CLASSIFY:
  GL Account:   {account['gl_account']}
  Description:  {account['description']}
  Functional Area: {account.get('functional_area', 'N/A')}
  Cost Element Type: {account.get('cost_element_type', 'N/A')}
  FY2023 Balance: ${account['posting_amount']:,.0f}
{similar_text}
Respond with JSON only."""


def strip_markdown(raw: str) -> str:
    raw = raw.strip()
    if raw.startswith('```'):
        parts = raw.split('```')
        inner = parts[1] if len(parts) >= 3 else parts[1] if len(parts) > 1 else raw
        inner = re.sub(r'^[a-zA-Z]*\n?', '', inner, count=1)
        return inner.strip()
    return raw


def classify_account(client: anthropic.Anthropic, account: dict, similar: list[dict]) -> dict | None:
    """Call Claude to classify one GL account. Returns parsed decision or None."""
    prompt = build_prompt(account, similar)

    for attempt in range(3):
        try:
            response = client.messages.create(
                model=MODEL,
                max_tokens=800,
                system=SYSTEM_PROMPT,
                messages=[{'role': 'user', 'content': prompt}],
                timeout=30.0,
            )
            raw = strip_markdown(response.content[0].text.strip())
            decision = json.loads(raw)

            # Validate
            if decision.get('suggested_category') not in VALID_CATEGORIES:
                raise ValueError(f"Bad category: {decision.get('suggested_category')}")
            if decision.get('suggested_caption') not in VALID_CAPTIONS:
                raise ValueError(f"Bad caption: {decision.get('suggested_caption')}")
            score = float(decision.get('confidence_score', 0))
            if not 0.0 <= score <= 1.0:
                raise ValueError(f"Bad score: {score}")

            return decision

        except anthropic.RateLimitError:
            log.warning(f"  Rate limited, waiting {5 * (attempt + 1)}s...")
            time.sleep(5 * (attempt + 1))
        except (json.JSONDecodeError, ValueError, KeyError) as e:
            log.warning(f"  Parse error (attempt {attempt + 1}): {e}")
            time.sleep(1)
        except Exception as e:
            log.error(f"  Unexpected error: {type(e).__name__}: {e}")
            return None

    return None


# ══════════════════════════════════════════════════════════════
# OUTPUT
# ══════════════════════════════════════════════════════════════

def save_results(results: list[dict], output_json: str, output_xlsx: str):
    """Save results to JSON and Excel."""
    # JSON
    with open(output_json, 'w') as f:
        json.dump(results, f, indent=2, default=str)
    log.info(f"Results saved to {output_json}")

    # Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DISE Mapping Results"

        # Header
        headers = [
            'GL Account', 'Description', 'Functional Area', 'FY2023 Balance',
            'DISE Category', 'Expense Caption', 'ASC Citation',
            'Confidence', 'Score', 'Reasoning',
        ]
        header_fill = PatternFill(start_color='0F1729', end_color='0F1729', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=10)
        thin_border = Border(
            bottom=Side(style='thin', color='E2E8F4'),
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left')

        # Category colors
        cat_colors = {
            'Purchases of inventory': 'DCFCE7',
            'Employee compensation': 'DBEAFE',
            'Depreciation': 'FEF3C7',
            'Intangible asset amortization': 'EDE9FE',
            'Other expenses': 'FEE2E2',
        }
        conf_colors = {
            'HIGH': '0A7C42',
            'MEDIUM': '92580A',
            'LOW': 'C81A1A',
        }

        for i, r in enumerate(results, 2):
            ws.cell(row=i, column=1, value=r['gl_account'])
            ws.cell(row=i, column=2, value=r['description'])
            ws.cell(row=i, column=3, value=r.get('functional_area', ''))
            ws.cell(row=i, column=4, value=r['posting_amount']).number_format = '#,##0'
            cat_cell = ws.cell(row=i, column=5, value=r.get('suggested_category', ''))
            cat_cell.fill = PatternFill(
                start_color=cat_colors.get(r.get('suggested_category', ''), 'FFFFFF'),
                end_color=cat_colors.get(r.get('suggested_category', ''), 'FFFFFF'),
                fill_type='solid',
            )
            ws.cell(row=i, column=6, value=r.get('suggested_caption', ''))
            ws.cell(row=i, column=7, value=r.get('suggested_citation', ''))
            conf_cell = ws.cell(row=i, column=8, value=r.get('confidence_label', ''))
            conf_cell.font = Font(
                color=conf_colors.get(r.get('confidence_label', ''), '000000'),
                bold=True,
            )
            ws.cell(row=i, column=9, value=r.get('confidence_score', 0)).number_format = '0%'
            ws.cell(row=i, column=10, value=r.get('draft_reasoning', ''))

            # Thin border on each row
            for col in range(1, 11):
                ws.cell(row=i, column=col).border = thin_border

        # Column widths
        widths = [12, 40, 22, 15, 30, 10, 22, 10, 8, 60]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col)].width = w

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Auto-filter
        ws.auto_filter.ref = f'A1:J{len(results) + 1}'

        wb.save(output_xlsx)
        log.info(f"Excel saved to {output_xlsx}")

    except Exception as e:
        log.error(f"Excel save failed: {e}")


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description='Offline DISE Mapping Agent')
    parser.add_argument('--batch', default='10', help='Number of accounts to classify (or "all")')
    parser.add_argument('--dry-run', action='store_true', help='Show prompts without calling Claude')
    parser.add_argument('--output', default='dise_mapping_results', help='Output file prefix')
    args = parser.parse_args()

    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key and not args.dry_run:
        log.error("Set ANTHROPIC_API_KEY environment variable")
        sys.exit(1)

    log.info("=" * 60)
    log.info("  BE TECHNOLOGY — OFFLINE DISE MAPPING AGENT")
    log.info("=" * 60)

    # Load data
    log.info("Loading GL accounts from DISE_Manual_Work_Demo.xlsx...")
    accounts = load_gl_accounts()
    log.info(f"  Loaded {len(accounts)} GL accounts")

    log.info("Loading reference mappings from DISE_Mapping_Toolkit.xlsx...")
    reference = load_reference_mappings()
    log.info(f"  Loaded {len(reference)} pre-mapped reference accounts")

    # Determine batch
    batch_size = len(accounts) if args.batch == 'all' else int(args.batch)
    batch_size = min(batch_size, len(accounts))
    to_process = accounts[:batch_size]
    log.info(f"  Processing {batch_size} of {len(accounts)} accounts")
    log.info(f"  Model: {MODEL}")
    log.info("")

    if not args.dry_run:
        client = anthropic.Anthropic(api_key=api_key)

    results = []
    stats = {'high': 0, 'medium': 0, 'low': 0, 'errors': 0}
    cat_counts = {c: 0 for c in VALID_CATEGORIES}
    start_time = time.time()

    for i, account in enumerate(to_process, 1):
        similar = find_similar(account['description'], reference)

        log.info(
            f"[{i}/{batch_size}] {account['gl_account']} — "
            f"{account['description'][:50]} "
            f"(${account['posting_amount']:,.0f})"
        )
        log.info(f"  Similar refs: {len(similar)} "
                 f"(best: {similar[0]['similarity_score']:.3f} '{similar[0]['description']}'" if similar else "(none)")

        if args.dry_run:
            prompt = build_prompt(account, similar)
            log.info(f"  [DRY RUN] Prompt length: {len(prompt)} chars")
            continue

        decision = classify_account(client, account, similar)

        if decision:
            result = {
                **account,
                **decision,
            }
            results.append(result)

            label = decision['confidence_label']
            stats[label.lower()] += 1
            cat_counts[decision['suggested_category']] += 1

            log.info(
                f"  -> {decision['suggested_category']} | "
                f"{decision['suggested_caption']} | "
                f"{label} ({decision['confidence_score']:.0%})"
            )
        else:
            log.error(f"  FAILED — could not classify")
            stats['errors'] += 1
            results.append({**account, 'error': 'Classification failed'})

        # Rate limiting
        if i < batch_size:
            time.sleep(API_DELAY)

    elapsed = time.time() - start_time

    if args.dry_run:
        log.info(f"\nDry run complete. {batch_size} prompts generated.")
        return

    # Summary
    log.info("")
    log.info("=" * 60)
    log.info("  CLASSIFICATION RESULTS")
    log.info("=" * 60)
    log.info(f"  Accounts processed: {batch_size}")
    log.info(f"  Time elapsed:       {elapsed:.1f}s ({elapsed/max(batch_size,1):.1f}s per account)")
    log.info(f"  Errors:             {stats['errors']}")
    log.info("")
    log.info("  Confidence distribution:")
    log.info(f"    HIGH:   {stats['high']}")
    log.info(f"    MEDIUM: {stats['medium']}")
    log.info(f"    LOW:    {stats['low']}")
    log.info("")
    log.info("  Category distribution:")
    for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
        pct = count / max(batch_size - stats['errors'], 1) * 100
        log.info(f"    {cat:35} {count:4}  ({pct:.0f}%)")
    log.info("=" * 60)

    # Save
    save_results(results, f'{args.output}.json', f'{args.output}.xlsx')

    log.info(f"\nDone. Open {args.output}.xlsx to review classifications.")


if __name__ == '__main__':
    main()
